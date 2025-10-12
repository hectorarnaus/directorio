import openpyxl 
import json
import os

wb = openpyxl.Workbook()
hoja = wb.active


fila=1


ruta=os.getcwd()+("/json")
for fichero_json in os.listdir(ruta):
# Abre el archivo y carga su contenido
    print(fichero_json)
    with open(os.path.join(ruta, fichero_json), "r", encoding="utf-8") as f:
        datos = json.load(f)

# Ahora 'datos' es una lista de diccionarios

    creperias = datos["creperias"]
    for negocio in creperias:
        hoja.cell(row=fila, column=1, value=negocio["nombre"])
        hoja.cell(row=fila, column=2, value=negocio["rating"])
        hoja.cell(row=fila, column=3, value=negocio["reviews"])
        hoja.cell(row=fila, column=4, value=negocio["horario"])
        hoja.cell(row=fila, column=5, value=negocio["direccion"])
        hoja.cell(row=fila, column=6, value=negocio["ciudad"])
        hoja.cell(row=fila, column=7, value=negocio["provincia"])
        hoja.cell(row=fila, column=8, value=negocio["web"])
        hoja.cell(row=fila, column=9, value=negocio["telefono"])
        hoja.cell(row=fila, column=10, value=negocio["mapa"])
        hoja.cell(row=fila, column=11, value=negocio["imagen"])
        hoja.cell(row=fila, column=12, value="\n".join(negocio["comentarios"]))
                                                    
        fila+=1


wb.save("datos.xlsx")