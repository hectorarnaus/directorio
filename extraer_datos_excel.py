import openpyxl 
from negocio import Negocio



def ultima_fila_real(hoja):
    # Recorre desde abajo hacia arriba buscando la primera fila con algún valor no vacío
    for i in range(1,hoja.max_row, 1):
        if hoja.cell(row=i,column=1).value==None:
           return i
    return hoja.max_row

def obten_lista_provincias(fichero_excel):
    lista=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<ultima_fila_real(hoja_activa):
            provincia=hoja_activa.cell(row=fila,column=5).value
            if provincia.isupper():
                provincia=provincia.capitalize()    
            if provincia not in lista:
                lista.append(provincia)
            fila+=1
        lista.sort()
        return lista
        
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def obten_lista_municipios(fichero_excel):
    lista=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<ultima_fila_real(hoja_activa):
            localidad=hoja_activa.cell(row=fila,column=4).value
            if localidad.isupper():
                localidad=localidad.capitalize()    
            if localidad not in lista:
                lista.append(localidad)
            fila+=1
        lista.sort()
        return lista

    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def obten_provincia_de_municipio(fichero_excel,municipio):
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<hoja_activa.max_row:
            if hoja_activa.cell(row=fila,column=7).value==municipio:
                return hoja_activa.cell(row=fila,column=8).value
            fila+=1
        return ""
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def limpiar_web(web):
    if web!=None:
        if web.find("?utm")!=-1:
            return web[:web.find("?utm")]
    return web


def obten_lista_negocios(fichero_excel):
    lista_negocios=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        while fila<ultima_fila_real(hoja_activa):
           
            nombre=hoja_activa.cell(row=fila,column=1).value if hoja_activa.cell(row=fila,column=1).value!=None else None
            direccion=hoja_activa.cell(row=fila,column=2).value if hoja_activa.cell(row=fila,column=2).value!=None else None
            CP=hoja_activa.cell(row=fila,column=3).value if hoja_activa.cell(row=fila,column=3).value!=None else None
            ciudad=hoja_activa.cell(row=fila,column=4).value if hoja_activa.cell(row=fila,column=4).value!=None else None
            provincia=hoja_activa.cell(row=fila,column=5).value if hoja_activa.cell(row=fila,column=5).value!=None else None
            telefono=hoja_activa.cell(row=fila,column=6).value if hoja_activa.cell(row=fila,column=6).value!=None else None
            pagina_web=limpiar_web(hoja_activa.cell(row=fila,column=7).value if hoja_activa.cell(row=fila,column=7).value!=None else None)
            actividad=hoja_activa.cell(row=fila,column=8).value if hoja_activa.cell(row=fila,column=8).value!=None else None
            actividades_relacionadas=hoja_activa.cell(row=fila,column=9).value if hoja_activa.cell(row=fila,column=9).value!=None else None
            marcas=hoja_activa.cell(row=fila,column=10).value if hoja_activa.cell(row=fila,column=10).value!=None else None
            descripcion=hoja_activa.cell(row=fila,column=11).value if hoja_activa.cell(row=fila,column=11).value!=None else None
            mapa=hoja_activa.cell(row=fila,column=12).value if hoja_activa.cell(row=fila,column=12).value!=None else None
            imagen=hoja_activa.cell(row=fila,column=13).value if hoja_activa.cell(row=fila,column=13).value!=None else None
            facebook=hoja_activa.cell(row=fila,column=14).value if hoja_activa.cell(row=fila,column=14).value!=None else None
            instagram=hoja_activa.cell(row=fila,column=15).value if hoja_activa.cell(row=fila,column=15).value!=None else None
            x=hoja_activa.cell(row=fila,column=16).value if hoja_activa.cell(row=fila,column=16).value!=None else None
            youtube=hoja_activa.cell(row=fila,column=17).value if hoja_activa.cell(row=fila,column=17).value!=None else None
            horario=hoja_activa.cell(row=fila,column=18).value if hoja_activa.cell(row=fila,column=18).value!=None else None
            descripcion_seo=hoja_activa.cell(row=fila,column=19).value  if hoja_activa.cell(row=fila,column=19).value!=None else None

            nuevo=Negocio(nombre,direccion,CP,ciudad,provincia,telefono,pagina_web,actividad,actividades_relacionadas,marcas,descripcion,mapa,imagen,facebook,instagram,x,youtube,horario,descripcion_seo)
            
            lista_negocios.append(nuevo)
            fila+=1
                  
        return lista_negocios
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")


def obten_lista_negocios_municipio(fichero_excel,municipio):
    lista_negocios=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        while fila<hoja_activa.max_row:
            if hoja_activa.cell(row=fila,column=4).value==municipio:
                nombre=hoja_activa.cell(row=fila,column=1).value if hoja_activa.cell(row=fila,column=1).value!=None else None
                direccion=hoja_activa.cell(row=fila,column=2).value if hoja_activa.cell(row=fila,column=2).value!=None else None
                CP=hoja_activa.cell(row=fila,column=3).value if hoja_activa.cell(row=fila,column=3).value!=None else None
                provincia=hoja_activa.cell(row=fila,column=5).value if hoja_activa.cell(row=fila,column=5).value!=None else None
                telefono=hoja_activa.cell(row=fila,column=6).value if hoja_activa.cell(row=fila,column=6).value!=None else None
                pagina_web=limpiar_web(hoja_activa.cell(row=fila,column=7).value if hoja_activa.cell(row=fila,column=7).value!=None else None)
                actividad=hoja_activa.cell(row=fila,column=8).value if hoja_activa.cell(row=fila,column=8).value!=None else None
                actividades_relacionadas=hoja_activa.cell(row=fila,column=9).value if hoja_activa.cell(row=fila,column=9).value!=None else None
                marcas=hoja_activa.cell(row=fila,column=10).value if hoja_activa.cell(row=fila,column=10).value!=None else None
                descripcion=hoja_activa.cell(row=fila,column=11).value if hoja_activa.cell(row=fila,column=11).value!=None else None
                mapa=hoja_activa.cell(row=fila,column=12).value if hoja_activa.cell(row=fila,column=12).value!=None else None
                imagen=hoja_activa.cell(row=fila,column=13).value if hoja_activa.cell(row=fila,column=13).value!=None else None
                facebook=hoja_activa.cell(row=fila,column=14).value if hoja_activa.cell(row=fila,column=14).value!=None else None
                instagram=hoja_activa.cell(row=fila,column=15).value if hoja_activa.cell(row=fila,column=15).value!=None else None
                x=hoja_activa.cell(row=fila,column=16).value if hoja_activa.cell(row=fila,column=16).value!=None else None
                youtube=hoja_activa.cell(row=fila,column=17).value if hoja_activa.cell(row=fila,column=17).value!=None else None
                horario=hoja_activa.cell(row=fila,column=18).value if hoja_activa.cell(row=fila,column=18).value!=None else None
                descripcion_seo=hoja_activa.cell(row=fila,column=19).value  if hoja_activa.cell(row=fila,column=19).value!=None else None

                nuevo=Negocio(nombre,direccion,CP,municipio,provincia,telefono,pagina_web,actividad,actividades_relacionadas,marcas,descripcion,mapa,imagen,facebook,instagram,x,youtube,horario,descripcion_seo)
                lista_negocios.append(nuevo)

            fila+=1
                  
        return lista_negocios
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")
               