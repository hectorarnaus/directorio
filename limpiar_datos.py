import openpyxl 
import os

def ultima_fila_con_datos(ws):
    """
    Devuelve el número de la última fila con datos en una hoja de openpyxl.
    Ignora filas vacías al final.
    """
    for fila in range(ws.max_row, 0, -1):
        if any(ws.cell(row=fila, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return fila
    return 0  # si la hoja está vacía


try:
    datos=openpyxl.load_workbook("datos.xlsx")
    hoja_activa = datos.active
    fila=1
    ultima=ultima_fila_con_datos(hoja_activa) + 1
    print("Última fila con datos: ",ultima)
    while fila<=ultima:
        print("Analizando fila ",fila)
        fila_analizada=fila+1
        igual=True
        while igual==True and fila_analizada<=ultima:
            print("  Comparando con fila ",fila_analizada)
            if (hoja_activa.cell(row=fila,column=1).value==hoja_activa.cell(row=fila_analizada,column=1).value and
                hoja_activa.cell(row=fila,column=5).value==hoja_activa.cell(row=fila_analizada,column=5).value and
                hoja_activa.cell(row=fila,column=6).value==hoja_activa.cell(row=fila_analizada,column=6).value and
                hoja_activa.cell(row=fila,column=7).value==hoja_activa.cell(row=fila_analizada,column=7).value):
                
                    hoja_activa.delete_rows(fila_analizada)
                    print("  Fila eliminada: ",fila_analizada)
                    fila_analizada+=1
            else:
                igual=False
        print(f"Pasando a fila {fila}")    
        
        fila += 1
    datos.save("datos_limpios.xlsx")
except FileNotFoundError:
    print("Error: Archivo no encontrado.")
except Exception as e:
    print(f"Ocurrió un error: {e}")