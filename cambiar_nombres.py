from os import *
from funciones_excel import ultima_fila_real
import openpyxl 
from re import sub

def normaliza_nombre(nombre):

    nombre = sub(r'-([a-z])', lambda m: '-' + m.group(1).upper(), nombre)
    nombre = sub(r"'([a-z])", lambda m: "'" + m.group(1).upper(), nombre)
    nombre=nombre.replace("D'","d'")
    nombre=nombre.replace("S'","s'")
    nombre=nombre.replace("L'","l'")
    nombre=nombre.replace(" De "," de ")
    nombre=nombre.replace(" Del "," del ")
    nombre=nombre.replace(" Y "," y ")
    nombre=nombre.replace(" I "," i ")
    return nombre        


def cambia_nombre_municipio_en_negocios(fichero_excel):
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        while fila<ultima_fila_real(hoja_activa):
           
            old=hoja_activa.cell(row=fila,column=4).value
            new=normaliza_nombre(hoja_activa.cell(row=fila,column=4).value)
            if old!=new:
                print(f"Nombre cambiado: {old} -> {new}")

            hoja_activa.cell(row=fila,column=4).value=normaliza_nombre(hoja_activa.cell(row=fila,column=4).value)
            fila+=1   
        datos.save(fichero_excel)    
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")   

def cambia_nombre_municipio_en_municipios(fichero_excel):
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        while fila<ultima_fila_real(hoja_activa):
           
            hoja_activa.cell(row=fila,column=1).value=normaliza_nombre(hoja_activa.cell(row=fila,column=1).value)
            fila+=1                    
        datos.save(fichero_excel)
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

ruta=getcwd()+("/directorio/provincia")
'''for archivo in listdir(ruta):
    nombre=archivo.split(".")[0]
    extension=archivo.split(".")[1]
    if nombre.find("-")!=-1 or nombre.find("'")!=-1:
        nombre=normaliza_nombre(nombre)
        print(f"Nombre cambiado: {nombre}")
        nuevo_nombre = f"{nombre}.{extension}"
        rename(ruta+"/"+archivo, ruta+"/"+nuevo_nombre)
'''
cambia_nombre_municipio_en_municipios(getcwd()+("/xslx/localidades.xlsx"))
cambia_nombre_municipio_en_negocios(getcwd()+("/xslx/empresas.xlsx"))